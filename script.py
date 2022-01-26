import os
import xlsxwriter
import openpyxl
import csv
import pandas as pd
import matplotlib.pyplot as plt

output_file = 'report.csv'
#values
start_row = 0
start_column = 0
max_commit = 0
commit_cell = 'A2'
commit_row = 1
author_cell = 'B2'
author_row = 1
date_cell = 'C2'
date_row = 1
#flags
commit_giulio = 0
commit_anastasia = 0
first = 1
#cursors
previous_date = ''
actual_date = ''

wb = xlsxwriter.Workbook('report.csv')
ws = wb.add_worksheet()
ws2 = wb.add_worksheet()

#chart object
chart = wb.add_chart({'type': 'line'})

#os.system('cmd /k "git log > report.txt"')

with open('report.txt', 'r') as txt:
	reader = csv.reader(txt, delimiter='\t')
	ws.write(start_row, start_column, 'Commit')
	ws.write(start_row, start_column+1, 'Author')
	ws.write(start_row, start_column+2, 'Date')
	#ws.append(['Commit', 'Author', 'Date'])
	ws2.write(start_row, start_column, 'Anastasia')
	ws2.write(start_row, start_column+1, 'Giulio')
	ws2.write(start_row, start_column+2, 'Date')
	#ws2.append(['Anastasia', 'Giulio', 'Date'])

	start_row = 1
	start_row_2 = start_row
	
	for row in reader:	
		if(row!=[]):
			res = row[0].split()
			for word in res:
				if(word=='commit'):
					max_commit += 1

					#inserting commit
					#ws[commit_cell]=res[1:][0]
					ws.write(start_row, start_column, res[1:][0])
					commit_row += 1
					commit_cell = 'A'+str(commit_row)
				if(word=='Author:'):

					#inserting author
					#ws[author_cell]=res[1:2][0]
					ws.write(start_row, start_column+1, res[1:2][0])
					author_row += 1
					author_cell = 'B'+str(author_row)
					if(res[1:2][0]=='Anastasia' or res[1:2][0]=='AnastasiaBrinati'):
						commit_anastasia+=1
					else:
						commit_giulio+=1
				if(word=='Date:'):

					#inserting date
					#ws[date_cell]=res[2:3][0]+res[5:6][0]
					ws.write(start_row, start_column+2, res[2:3][0]+res[5:6][0])
					actual_date = res[2:3][0]+res[5:6][0]
					date_row += 1
					date_cell = 'C'+str(date_row)
					if(previous_date!=actual_date):
						if(first!=1):
							#ws2.append([str(commit_anastasia),str(commit_giulio), previous_date])
							ws2.write(start_row_2, start_column, str(commit_anastasia))
							ws2.write(start_row_2, start_column+1, str(commit_giulio))
							ws2.write(start_row_2, start_column+2, previous_date)
							start_row_2+=1
							
							commit_anastasia = 0
							commit_giulio = 0
					previous_date = actual_date
					first+=1
					start_row+=1
#ws2.append([str(commit_anastasia),str(commit_giulio), previous_date])
ws2.write(start_row_2, start_column, str(commit_anastasia))
ws2.write(start_row_2, start_column+1, str(commit_giulio))
ws2.write(start_row_2, start_column+2, previous_date)

# Add a chart title.
chart.set_title({'name': 'Process Control Chart: commits'})
# Add x-axis label
chart.set_x_axis({
	'name': 'Date',
})
# Add y-axis label
chart.set_y_axis({
	'name': 'Commits',
	'min' : 0,
	'max' : 100,
	'interval_tick': 3
})
# Set an Excel chart style.
chart.set_style(11)

#giulio
chart.add_series({
    'categories': '=Sheet2!$C$2:$C$4',
    'values':     '=Sheet2!$B$2:$A$3',
    'line':       {'color': 'blue'}
})

#anastasia
chart.add_series({
    'categories': '=Sheet2!$C$2:$C$4',
    'values':     '=Sheet2!$A$2:$A$4'
})

#media
chart.add_series({
    'categories': '=Sheet2!$C$2:$C$4',
    'values':     '50',
    'line':       {'color': 'green'}
})

#upperbound
chart.add_series({
    'categories': '=Sheet2!$C$2:$C$4',
    'values':     '90',
    'line':       {'color': 'red'}
})

#lowerbound
chart.add_series({
    'categories': '=Sheet2!$C$2:$C$4',
    'values':     '5',
    'line':       {'color': 'red'}
})
 
# add chart to the worksheet with given
# offset values at the top-left corner of
# a chart is anchored to cell D2
ws2.insert_chart('D2', chart, {'x_offset': 20, 'y_offset': 5})

#wb.save(output_file)
print(max_commit)
wb.close()