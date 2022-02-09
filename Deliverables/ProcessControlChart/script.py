import os
import xlsxwriter
import openpyxl
import csv
import pandas as pd
import matplotlib.pyplot as plt
import subprocess
import sys
import time

output_file = 'report.csv'
#values
start_row = 0
start_column = 0
max_commit = 0
commit_cell = 'A2'
commit_row = 1
author_cell = 'B2'
author_row = 1
date_cell = 'C2'
date_row = 1
#flags
commit_giulio = 0
commit_anastasia = 0
first = 1
#cursors
previous_date = ''
actual_date = ''

average = 0


wb = xlsxwriter.Workbook('report.csv')
ws = wb.add_worksheet()
ws2 = wb.add_worksheet()

#chart object
chart = wb.add_chart({'type': 'line'})

#os.system('cmd /k "git log > report.txt"')

with open('report.txt', 'r') as txt:
	reader = csv.reader(txt, delimiter='\t')
	ws.write(start_row, start_column, 'Commit')
	ws.write(start_row, start_column+1, 'Author')
	ws.write(start_row, start_column+2, 'Date')
	#ws.append(['Commit', 'Author', 'Date'])
	ws2.write(start_row, start_column, 'Date')
	ws2.write(start_row, start_column+1, 'Giulio')
	ws2.write(start_row, start_column+2, 'Anastasia')

	ws2.write(start_row, start_column+3, 'LCL')
	ws2.write(start_row, start_column+4, 'Average')
	ws2.write(start_row, start_column+5, 'UCL')
	#ws2.append(['Anastasia', 'Giulio', 'Date'])

	start_row = 1
	start_row_2 = start_row
	
	for row in reader:	
		if(row!=[]):
			res = row[0].split()
			for word in res:
				if(word=='commit'):
					max_commit += 1

					#inserting commit
					#ws[commit_cell]=res[1:][0]
					ws.write(start_row, start_column, res[1:][0])
					commit_row += 1
					commit_cell = 'A'+str(commit_row)
				if(word=='Author:'):

					#inserting author
					#ws[author_cell]=res[1:2][0]
					ws.write(start_row, start_column+1, res[1:2][0])
					author_row += 1
					author_cell = 'B'+str(author_row)
					if(res[1:2][0]=='Anastasia' or res[1:2][0]=='AnastasiaBrinati'):
						commit_anastasia+=1
					else:
						commit_giulio+=1
				if(word=='Date:'):

					#inserting date
					#ws[date_cell]=res[2:3][0]+res[5:6][0]
					ws.write(start_row, start_column+2, res[2:3][0]+res[5:6][0])
					actual_date = res[2:3][0]+res[5:6][0]
					date_row += 1
					date_cell = 'C'+str(date_row)
					if((previous_date!=actual_date) and (first!=1)):
							#ws2.append([str(commit_anastasia),str(commit_giulio), previous_date])
							ws2.write(start_row_2, start_column+2, commit_anastasia)
							ws2.write(start_row_2, start_column+1, commit_giulio)
							ws2.write(start_row_2, start_column, previous_date)
							start_row_2+=1

							average += (commit_anastasia + commit_giulio)
							
							commit_anastasia = 0
							commit_giulio = 0
					previous_date = actual_date
					first+=1
					start_row+=1
#ws2.append([str(commit_anastasia),str(commit_giulio), previous_date])
ws2.write(start_row_2, start_column+2, commit_anastasia)
ws2.write(start_row_2, start_column+1, commit_giulio)
ws2.write(start_row_2, start_column, previous_date)
start_row_2+=1

average = average/(start_row_2*2)

for i in range(1, start_row_2):
	ws2.write(i, start_column+3, 10)
	ws2.write(i, start_column+4, average)
	ws2.write(i, start_column+5, 90)

# Add a chart title.
chart.set_title({'name': 'Process Control Chart: commits'})
# Add x-axis label
chart.set_x_axis({
	'name': 'Date',
})
# Add y-axis label
chart.set_y_axis({
	'name': 'Commits',
	'min' : 0,
	'max' : 100,
	'interval_tick': 3
})
# Set an Excel chart style.
chart.set_style(11)

#giulio
chart.add_series({
    'categories': ['Sheet2', 1, 0, start_row_2, 0],
    'values':     ['Sheet2', 1, 1, start_row_2, 1],
    'line':       {'color': 'blue'},
    'name': 'Giulio'
})

#anastasia
chart.add_series({
    'categories': ['Sheet2', 1, 0, start_row_2, 0],
    'values':     ['Sheet2', 1, 2, start_row_2, 2],
    'name': 'Anastasia'
})

#media
chart.add_series({
    'categories': ['Sheet2', 1, 0, start_row_2, 0],
    'values':     ['Sheet2', 1, 4, start_row_2, 4],
    'line':       {'color': 'green'},
    'name': 'Average'
})

#upperbound
chart.add_series({
    'categories': ['Sheet2', 1, 0, start_row_2, 0],
    'values':     ['Sheet2', 1, 5, start_row_2, 5],
    'line':       {'color': 'red'},
    'name': 'UCL'
})

#lowerbound
chart.add_series({
    'categories': ['Sheet2', 1, 0, start_row_2, 0],
    'values':     ['Sheet2', 1, 3, start_row_2, 3],
    'line':       {'color': 'red'},
    'name': 'LCL'
})
 
# add chart to the worksheet with given
# offset values at the top-left corner of
# a chart is anchored to cell D2
ws2.insert_chart('G2', chart, {'x_offset': 20, 'y_offset': 5})

#wb.save(output_file)
wb.close()